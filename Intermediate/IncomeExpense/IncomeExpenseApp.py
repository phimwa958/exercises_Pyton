import wx
import os
from datetime import datetime
import openpyxl

class IncomeExpenseApp(wx.Frame):
    def __init__(self, *args, **kwargs):
        super(IncomeExpenseApp, self).__init__(*args, **kwargs)

        self.data_file = "income_expense_data.xlsx"
        self.load_data()

        self.InitUI()

    def InitUI(self):
        panel = wx.Panel(self)

        vbox = wx.BoxSizer(wx.VERTICAL)

        # Title
        title = wx.StaticText(panel, label="Income and Expense Tracker")
        title_font = wx.Font(14, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        title.SetFont(title_font)
        vbox.Add(title, flag=wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, border=10)

        # Input fields
        hbox1 = wx.BoxSizer(wx.HORIZONTAL)
        hbox1.Add(wx.StaticText(panel, label="Description:"), flag=wx.RIGHT, border=8)
        self.description_txt = wx.TextCtrl(panel)
        hbox1.Add(self.description_txt, proportion=1)
        vbox.Add(hbox1, flag=wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        hbox2 = wx.BoxSizer(wx.HORIZONTAL)
        hbox2.Add(wx.StaticText(panel, label="Amount:"), flag=wx.RIGHT, border=8)
        self.amount_txt = wx.TextCtrl(panel)
        hbox2.Add(self.amount_txt, proportion=1)
        vbox.Add(hbox2, flag=wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        hbox3 = wx.BoxSizer(wx.HORIZONTAL)
        hbox3.Add(wx.StaticText(panel, label="Notes:"), flag=wx.RIGHT, border=8)
        self.notes_txt = wx.TextCtrl(panel)
        hbox3.Add(self.notes_txt, proportion=1)
        vbox.Add(hbox3, flag=wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        # Buttons
        hbox4 = wx.BoxSizer(wx.HORIZONTAL)
        self.income_btn = wx.Button(panel, label="Add Income")
        self.expense_btn = wx.Button(panel, label="Add Expense")
        hbox4.Add(self.income_btn, flag=wx.RIGHT, border=5)
        hbox4.Add(self.expense_btn)
        vbox.Add(hbox4, flag=wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, border=10)

        self.income_btn.Bind(wx.EVT_BUTTON, self.on_add_income)
        self.expense_btn.Bind(wx.EVT_BUTTON, self.on_add_expense)

        # Display area
        self.list_ctrl = wx.ListCtrl(panel, style=wx.LC_REPORT)
        self.list_ctrl.InsertColumn(0, "Date", width=150)
        self.list_ctrl.InsertColumn(1, "Description", width=150)
        self.list_ctrl.InsertColumn(2, "Amount", width=100)
        self.list_ctrl.InsertColumn(3, "Type", width=100)
        self.list_ctrl.InsertColumn(4, "Notes", width=200)
        vbox.Add(self.list_ctrl, proportion=1, flag=wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, border=10)

        # Total
        hbox5 = wx.BoxSizer(wx.HORIZONTAL)
        self.total_lbl = wx.StaticText(panel, label="Total Balance: 0")
        hbox5.Add(self.total_lbl, flag=wx.ALIGN_LEFT)
        vbox.Add(hbox5, flag=wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP | wx.BOTTOM, border=10)

        # Button to read the balance
        read_balance_btn = wx.Button(panel, label="Show Balance")
        vbox.Add(read_balance_btn, flag=wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, border=10)
        read_balance_btn.Bind(wx.EVT_BUTTON, self.on_show_balance)

        panel.SetSizer(vbox)

        self.SetTitle("Income and Expense Tracker")
        self.SetSize((700, 600))
        self.Centre()
        self.update_display()

    def on_add_income(self, event):
        self.add_entry("Income")

    def on_add_expense(self, event):
        self.add_entry("Expense")

    def add_entry(self, entry_type):
        description = self.description_txt.GetValue()
        notes = self.notes_txt.GetValue()
        try:
            amount = float(self.amount_txt.GetValue())
        except ValueError:
            wx.MessageBox("Please enter a valid amount", "Error", wx.OK | wx.ICON_ERROR)
            return

        if not description:
            description = "(No description)"

        entry = {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "description": description,
            "amount": amount,
            "type": entry_type,
            "notes": notes
        }
        
        self.save_data(entry)
        self.update_display()

        self.description_txt.SetValue("")
        self.amount_txt.SetValue("")
        self.notes_txt.SetValue("")

    def update_display(self):
        self.list_ctrl.DeleteAllItems()
        total_balance = 0

        workbook = openpyxl.load_workbook(self.data_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            date, description, amount, entry_type, notes = row
            index = self.list_ctrl.InsertItem(self.list_ctrl.GetItemCount(), date)
            self.list_ctrl.SetItem(index, 1, description)
            self.list_ctrl.SetItem(index, 2, f"{amount:.2f}")
            self.list_ctrl.SetItem(index, 3, entry_type)
            self.list_ctrl.SetItem(index, 4, notes or "")

            if entry_type == "Income":
                total_balance += amount
            elif entry_type == "Expense":
                total_balance -= amount

        self.total_lbl.SetLabel(f"Total Balance: {total_balance:.2f}")
        workbook.close()

    def on_show_balance(self, event):
        total_balance = self.total_lbl.GetLabel()
        wx.MessageBox(total_balance, "Current Balance", wx.OK | wx.ICON_INFORMATION)

    def load_data(self):
        if not os.path.exists(self.data_file):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "IncomeExpense"
            sheet.append(["Date", "Description", "Amount", "Type", "Notes"])
            workbook.save(self.data_file)

    def save_data(self, entry):
        workbook = openpyxl.load_workbook(self.data_file)
        sheet = workbook.active
        sheet.append([entry["date"], entry["description"], entry["amount"], entry["type"], entry["notes"]])
        workbook.save(self.data_file)

if __name__ == "__main__":
    app = wx.App()
    frame = IncomeExpenseApp(None)
    frame.Show()
    app.MainLoop()
